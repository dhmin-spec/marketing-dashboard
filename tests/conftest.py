# tests/conftest.py
import io
import openpyxl
import pytest

SHEET_TND = "검색광고 T&D"
SHEET_UPLOAD = "업로드용"


def _build() -> bytes:
    wb = openpyxl.Workbook()
    # sheet 1: 드롭다운 수식 (반드시 무변경으로 남아야 함)
    d = wb.active
    d.title = "드롭다운 수식"
    d["B1"] = "가나보험"
    d["C1"] = "=B1"  # 수식 보존 확인용

    # sheet 2: 검색광고 T&D
    t = wb.create_sheet(SHEET_TND)
    t["B4"], t["C4"], t["D4"] = "보종", "NO", "문구 및 이미지"
    t["E4"], t["F4"], t["H4"] = "글자수", "광고위치", "비고"
    rows = [
        (1, "판매 수수료가 없어 저렴한 가나다이렉트", "설명문구", 45),
        (2, "AI맞춤보장", "추가제목", 15),
        (3, "보험료확인", "서브링크", 6),
    ]
    for i, (no, text, pos, mx) in enumerate(rows):
        r = 5 + i
        t.cell(r, 2, "전 보종")
        t.cell(r, 3, no)
        t.cell(r, 4, text)
        t.cell(r, 5, f"=LEN(D{r})")
        t.cell(r, 6, pos)
        t.cell(r, 8, mx)

    # sheet 3: 업로드용 (C=문안 고정값, D=수식)
    u = wb.create_sheet(SHEET_UPLOAD)
    for i, (_no, text, _pos, _mx) in enumerate(rows):
        r = 1 + i
        u.cell(r, 1, "전 보종")
        u.cell(r, 2, ">")
        u.cell(r, 3, text)
        u.cell(r, 4, f"=A{r}&B{r}&C{r}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def sample_workbook_bytes() -> bytes:
    return _build()
