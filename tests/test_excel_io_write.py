import io
import openpyxl
import pytest
from shim.excel_io import build_revised_workbook, read_copies, SheetNotFoundError


def _load(data, keep_formulas=True):
    return openpyxl.load_workbook(io.BytesIO(data), data_only=not keep_formulas)


def test_overwrites_only_targeted_copy(sample_workbook_bytes):
    out = build_revised_workbook(sample_workbook_bytes, {2: "AI맞춤보장보험"})
    wb = _load(out)
    t = wb["검색광고 T&D"]
    assert t["D5"].value == "판매 수수료가 없어 저렴한 가나다이렉트"  # 미수정 유지
    assert t["D6"].value == "AI맞춤보장보험"                              # 수정 반영


def test_e_column_formula_preserved(sample_workbook_bytes):
    out = build_revised_workbook(sample_workbook_bytes, {2: "AI맞춤보장보험"})
    t = _load(out)["검색광고 T&D"]
    assert t["E6"].value == "=LEN(D6)"  # 글자수 수식 그대로


def test_upload_sheet_c_synced_d_formula_kept(sample_workbook_bytes):
    out = build_revised_workbook(sample_workbook_bytes, {2: "AI맞춤보장보험"})
    u = _load(out)["업로드용"]
    assert u["C2"].value == "AI맞춤보장보험"   # NO 2 -> 업로드용 2행
    assert u["D2"].value == "=A2&B2&C2"        # 연결 수식 유지


def test_other_sheet_untouched(sample_workbook_bytes):
    out = build_revised_workbook(sample_workbook_bytes, {2: "AI맞춤보장보험"})
    d = _load(out)["드롭다운 수식"]
    assert d["B1"].value == "가나보험"
    assert d["C1"].value == "=B1"


def test_missing_tnd_sheet_raises():
    wb = openpyxl.Workbook()
    wb.active.title = "드롭다운 수식"
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    with pytest.raises(SheetNotFoundError):
        build_revised_workbook(data, {1: "x"})


def test_upload_row_out_of_range_skipped():
    SHEET_TND = "검색광고 T&D"
    SHEET_UPLOAD = "업로드용"

    wb = openpyxl.Workbook()
    d = wb.active
    d.title = "드롭다운 수식"

    t = wb.create_sheet(SHEET_TND)
    t["B4"], t["C4"], t["D4"] = "보종", "NO", "문구 및 이미지"
    t["E4"], t["F4"], t["H4"] = "글자수", "광고위치", "비고"
    rows = [
        (1, "판매 수수료가 없어 저렴한 가나다이렉트", "설명문구", 45),
        (2, "AI맞춤보장", "추가제목", 15),
    ]
    for i, (no, text, pos, mx) in enumerate(rows):
        r = 5 + i
        t.cell(r, 2, "전 보종")
        t.cell(r, 3, no)
        t.cell(r, 4, text)
        t.cell(r, 5, f"=LEN(D{r})")
        t.cell(r, 6, pos)
        t.cell(r, 8, mx)

    # 업로드용에는 NO 1 한 행만 존재
    u = wb.create_sheet(SHEET_UPLOAD)
    u.cell(1, 1, "전 보종")
    u.cell(1, 2, ">")
    u.cell(1, 3, rows[0][1])
    u.cell(1, 4, "=A1&B1&C1")

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    out = build_revised_workbook(data, {2: "수정됨"})

    wb_out = _load(out)
    t_out = wb_out[SHEET_TND]
    assert t_out["D6"].value == "수정됨"  # NO 2, row 6, T&D는 정상 반영

    u_out = wb_out[SHEET_UPLOAD]
    assert u_out.max_row == 1  # phantom row 생성되지 않음


def test_formula_no_without_cache_is_skipped():
    # openpyxl로 생성한 파일은 수식 셀의 캐시값(cached value)을 담지 못한다.
    # 실제 엑셀에서 저장한 파일은 수식 계산 결과가 캐시로 함께 저장되므로
    # read_copies가 data_only=True 로드에서 그 캐시값을 읽어 정상 처리한다.
    # 여기서는 그 반대 상황(캐시 없음)을 재현해 해당 행이 스킵됨을 문서화한다.
    SHEET_TND = "검색광고 T&D"

    wb = openpyxl.Workbook()
    d = wb.active
    d.title = "드롭다운 수식"

    t = wb.create_sheet(SHEET_TND)
    t["B4"], t["C4"], t["D4"] = "보종", "NO", "문구 및 이미지"
    t["E4"], t["F4"], t["H4"] = "글자수", "광고위치", "비고"

    r = 5
    t.cell(r, 2, "전 보종")
    t.cell(r, 3, "=1")  # NO가 수식 — 캐시값 없이 openpyxl로 저장됨
    t.cell(r, 4, "문구")
    t.cell(r, 5, f"=LEN(D{r})")
    t.cell(r, 6, "설명문구")
    t.cell(r, 8, 45)

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    copies = read_copies(data)
    assert len(copies) == 0  # 캐시값이 없어 NO를 해석하지 못해 스킵됨
