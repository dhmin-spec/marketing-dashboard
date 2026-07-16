import io
from dataclasses import dataclass
import openpyxl

SHEET_TND = "검색광고 T&D"
SHEET_UPLOAD = "업로드용"
HEADER_ROW = 4
DATA_START = 5
COL_NO = 3
COL_COPY = 4
COL_POS = 6
COL_MAX = 8


class SheetNotFoundError(Exception):
    pass


@dataclass
class Copy:
    no: int
    text: str
    position: str
    max_len: int | None
    row: int


def _resolve_int(raw, cached) -> int | None:
    """NO 컬럼이 수식(예: =C5+1)인 경우 캐시된 계산값(cached)을 사용."""
    if isinstance(raw, str) and raw.startswith("="):
        raw = cached
    try:
        return int(raw)
    except (ValueError, TypeError):
        return None


def read_copies(data: bytes) -> list[Copy]:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    wb_vals = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    if SHEET_TND not in wb.sheetnames:
        raise SheetNotFoundError(f"시트 '{SHEET_TND}' 없음. 있는 시트: {wb.sheetnames}")
    ws = wb[SHEET_TND]
    ws_vals = wb_vals[SHEET_TND]
    copies: list[Copy] = []
    for r in range(DATA_START, ws.max_row + 1):
        no = ws.cell(r, COL_NO).value
        text = ws.cell(r, COL_COPY).value
        if no is None or text is None:
            continue
        no_int = _resolve_int(no, ws_vals.cell(r, COL_NO).value)
        if no_int is None:
            continue
        mx = ws.cell(r, COL_MAX).value
        copies.append(Copy(
            no=no_int,
            text=str(text),
            position=str(ws.cell(r, COL_POS).value or ""),
            max_len=int(mx) if isinstance(mx, (int, float)) else None,
            row=r,
        ))
    return copies


def build_revised_workbook(data: bytes, revisions: dict[int, str]) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    wb_vals = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    if SHEET_TND not in wb.sheetnames:
        raise SheetNotFoundError(f"시트 '{SHEET_TND}' 없음. 있는 시트: {wb.sheetnames}")
    ws = wb[SHEET_TND]
    ws_vals = wb_vals[SHEET_TND]

    order: list[int] = []  # T&D 데이터 순서대로의 NO 목록
    for r in range(DATA_START, ws.max_row + 1):
        no = ws.cell(r, COL_NO).value
        text = ws.cell(r, COL_COPY).value
        if no is None or text is None:
            continue
        no_int = _resolve_int(no, ws_vals.cell(r, COL_NO).value)
        if no_int is None:
            continue
        order.append(no_int)
        if no_int in revisions:
            ws.cell(r, COL_COPY).value = revisions[no_int]

    if SHEET_UPLOAD in wb.sheetnames:
        us = wb[SHEET_UPLOAD]
        for idx, no in enumerate(order):
            if no in revisions and idx + 1 <= us.max_row:
                us.cell(idx + 1, 3).value = revisions[no]  # C열

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
